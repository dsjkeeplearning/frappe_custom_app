# Copyright (c) 2025
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import openpyxl
from openpyxl.styles import Font, Alignment
from frappe.utils import getdate, add_days
from frappe import _
import json


class AttendanceExcelGenerator(Document):
    pass


# ---------------------------------------
# LEAVE CODE MAPPING
# ---------------------------------------
def map_leave_code(leave_type):
    if not leave_type:
        return ""

    leave_type = leave_type.lower()

    if leave_type.startswith("casual"):
        return "CL"
    elif leave_type.startswith("sick"):
        return "SL"
    elif leave_type.startswith("earn"):
        return "EL"
    elif leave_type.startswith("option") or leave_type.startswith("optional"):
        return "OH"
    elif leave_type.startswith("pat"):
        return "PTRL"
    elif leave_type.startswith("mat"):
        return "MTRL"
    elif leave_type.startswith("leave without pay") or leave_type.startswith("lop"):
        return "LOP"
    else:
        return "L"


# ---------------------------------------
# HOLIDAY CHECK FUNCTION
# ---------------------------------------
def get_holiday_code(employee, date):
    """Return WO (week off) or PH (public holiday) based on employee holiday list."""
    
    holiday_list = frappe.db.get_value("Employee", employee, "holiday_list")
    if not holiday_list:
        return None

    holiday = frappe.get_all(
        "Holiday",
        filters={"parent": holiday_list, "holiday_date": date},
        fields=["description"],
        limit=1
    )

    if not holiday:
        return None

    desc = (holiday[0].description or "").lower()

    if "saturday" in desc or "sunday" in desc:
        return "WO"

    return "PH"


# ---------------------------------------
# FETCH ATTENDANCE REQUEST FROM ATTENDANCE
# ---------------------------------------
def get_attendance_request(attendance_doc):
    """Returns request_type if linked attendance_request exists."""
    request_id = attendance_doc.get("attendance_request")
    if not request_id:
        return None

    req = frappe.get_value(
        "Attendance Request",
        request_id,
        ["reason"],
        as_dict=True
    )

    if not req:
        return None

    return (req.reason or "").lower()


# ---------------------------------------
# MAIN FUNCTION
# ---------------------------------------
@frappe.whitelist()
def generate_excel(doc):

    if isinstance(doc, str):
        doc = json.loads(doc)

    doc = frappe._dict(doc)

    if not (doc.company and doc.from_date and doc.to_date):
        frappe.throw(_("Please select Company, From Date and To Date"))

    employees = frappe.get_all(
        "Employee",
        filters={"company": doc.company},
        fields=["name", "employee", "employee_name", "employee_number"]
    )

    start = getdate(doc.from_date)
    end = getdate(doc.to_date)

    dates = []
    d = start
    while d <= end:
        dates.append(d)
        d = add_days(d, 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # HEADER
    ws.cell(row=1, column=1, value="EmpCode").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Short Name").font = Font(bold=True)
    ws.cell(row=1, column=3, value="FromDate").font = Font(bold=True)
    ws.cell(row=1, column=4, value="ToDate").font = Font(bold=True)

    formatted_from_date = start.strftime("%d/%m/%Y")
    formatted_to_date = end.strftime("%d/%m/%Y")

    for i, date in enumerate(dates):
        col = i + 5
        ws.cell(row=1, column=col, value=date.strftime("%d/%m/%Y"))
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    # BODY ROWS
    row = 2
    for emp in employees:

        ws.cell(row=row, column=1, value=emp.employee_number or emp.name)
        ws.cell(row=row, column=2, value=emp.employee_name)
        ws.cell(row=row, column=3, value=formatted_from_date)
        ws.cell(row=row, column=4, value=formatted_to_date)

        for i, date in enumerate(dates):
            col = i + 5

            attendance = frappe.get_value(
                "Attendance",
                {
                    "employee": emp.name,
                    "attendance_date": date,
                    "docstatus": 1
                },
                ["name", "status", "leave_type", "half_day_status", "attendance_request"],
                as_dict=True
            )

            value = ""
            holiday_code = get_holiday_code(emp.name, date)

            if attendance:
                status = (attendance.status or "").lower()
                leave_type = (attendance.leave_type or "")
                half_day_status = (attendance.half_day_status or "").lower()
                leave_code = map_leave_code(leave_type)

                # CHECK ATTENDANCE REQUEST
                req_type = get_attendance_request(attendance)

                # ---------------------
                # HALF DAY
                # ---------------------
                if status == "half day":

                    if half_day_status == "present":
                        if leave_code:
                            base = f"HD{leave_code}"
                        else:
                            base = "HD"

                    elif half_day_status == "absent":
                        if leave_code:
                            base = f"HD{leave_code}"
                        else:
                            base = "HD"

                    # Apply (OD)
                    if req_type == "on duty":
                        base = f"{base}(OD)"

                    if holiday_code:
                        value = f"{base},{holiday_code}"
                    else:
                        value = base

                # ---------------------
                # PRESENT
                # ---------------------
                elif status == "present":
                    if req_type == "on duty":
                        base = "P(OD)"
                    else:
                        base = "P"

                    if holiday_code:
                        value = f"{base},{holiday_code}"
                    else:
                        value = base

                # ---------------------
                # ABSENT
                # ---------------------
                elif status == "absent":
                    base = "A"
                    if holiday_code:
                        value = f"A,{holiday_code}"
                    else:
                        value = base

                # ---------------------
                # ON LEAVE
                # ---------------------
                elif status == "on leave":
                    base = leave_code or "L"
                    if holiday_code:
                        value = f"{base},{holiday_code}"
                    else:
                        value = base

                # ---------------------
                # WORK FROM HOME
                # ---------------------
                elif status == "work from home":
                    base = "WFH"
                    if holiday_code:
                        value = f"WFH,{holiday_code}"
                    else:
                        value = base

            else:
                if holiday_code:
                    value = holiday_code
                else:
                    value = ""

            ws.cell(row=row, column=col, value=value)

        row += 1

    filename = f"Attendance-{doc.company}-{doc.from_date}-to-{doc.to_date}.xlsx"
    filepath = frappe.utils.get_site_path("public", "files", filename)
    wb.save(filepath)

    return f"/files/{filename}"
