import frappe
from frappe import _
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from frappe.utils import nowdate


@frappe.whitelist()
def download_budget_excel(company):
    if not company:
        frappe.throw(_("Company is required"))

    # Fetch Expense Accounts
    accounts = frappe.get_all(
        "Account",
        filters={
            "company": company,
            "root_type": "Expense",
            "disabled": 0,
            "is_group": 0
        },
        fields=["name"],
        order_by="name"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Format"

    # Months from April to March
    months = [
        "April", "May", "June", "July", "August", "September",
        "October", "November", "December", "January", "February", "March"
    ]

    # Header Row
    ws.cell(row=1, column=1, value="Acc Name")
    for col, month in enumerate(months, start=2):
        ws.cell(row=1, column=col, value=month)

    # Account Rows
    row = 2
    for acc in accounts:
        ws.cell(row=row, column=1, value=acc.name)
        row += 1

    # Auto column width
    for col in range(1, len(months) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Send file response
    frappe.response.filename = f"Budget_Format_{company}.xlsx"
    frappe.response.filecontent = output.read()
    frappe.response.type = "binary"